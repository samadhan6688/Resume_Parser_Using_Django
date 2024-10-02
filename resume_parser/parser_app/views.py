import os
import re
from django.shortcuts import render
from PyPDF2 import PdfReader
from openpyxl import Workbook, load_workbook
from .forms import ResumeUploadForm
from .models import ResumeData

def parse_resume_pdf(pdf_file):
    pdf_reader = PdfReader(pdf_file)
    text = ''
    for page in pdf_reader.pages:
        text += page.extract_text() or ''

    # Log the extracted text for debugging
    print("Extracted text:", text)

    name = extract_name(text)
    email = extract_email(text)
    mobile_number = extract_mobile_number(text)
    linkedin_link = extract_linkedin_link(text)
    education = extract_education(text)
    experience = extract_experience(text)
    skills = extract_skills(text)

    return {
        'name': name,
        'email': email,
        'mobile_number': mobile_number,
        'linkedin_link': linkedin_link,
        'education': education,
        'experience': experience,
        'skills': skills,
    }

def extract_name(text):
    lines = text.splitlines()
    # Assume name is the first line; this can be modified based on your requirements
    return lines[0].strip() if lines else 'Unknown'

def extract_email(text):
    email_regex = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    match = re.search(email_regex, text)
    return match.group(0) if match else 'Not provided'

def extract_mobile_number(text):
    phone_regex = r'\+?[0-9]*[\s.-]?[0-9]+'
    match = re.search(phone_regex, text)
    return match.group(0) if match else 'Not provided'

def extract_linkedin_link(text):
    linkedin_regex = r'https?://(www\.)?linkedin\.com/in/[A-Za-z0-9-_]+'
    match = re.search(linkedin_regex, text)
    return match.group(0) if match else 'Not provided'

def extract_education(text):
    # Example: looking for keywords like "Education" or "Degree"
    education_keywords = ['Education', 'Degree', 'University', 'School']
    for line in text.splitlines():
        if any(keyword in line for keyword in education_keywords):
            return line.strip()
    return 'Not provided'

def extract_experience(text):
    # Example: looking for keywords like "Experience" or "Work"
    experience_keywords = ['Experience', 'Work', 'Employment']
    for line in text.splitlines():
        if any(keyword in line for keyword in experience_keywords):
            return line.strip()
    return 'Not provided'

def extract_skills(text):
    # Example: looking for keywords like "Skills"
    skills_keywords = ['Skills', 'Technical Skills']
    for line in text.splitlines():
        if any(keyword in line for keyword in skills_keywords):
            return line.strip()
    return 'Not provided'

def upload_resume(request):
    if request.method == 'POST':
        form = ResumeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            pdf_file = request.FILES['pdf_file']
            resume_data = parse_resume_pdf(pdf_file)

            # Save to database
            ResumeData.objects.create(**resume_data)

            # Define file path and create directory if it doesn't exist
            directory = 'uploads'  # Change this to 'data' if needed
            file_path = os.path.join(directory, 'resume_data.xlsx')
            if not os.path.exists(directory):
                os.makedirs(directory)  # Create the directory if it doesn't exist

            # Check if the Excel file exists
            if not os.path.exists(file_path):
                workbook = Workbook()
                sheet = workbook.active

                # Write headers
                headers = ['Name', 'Email', 'Mobile Number', 'LinkedIn', 'Education', 'Experience', 'Skills']
                sheet.append(headers)
                workbook.save(file_path)  # Save the new workbook with headers
            else:
                workbook = load_workbook(file_path)
                sheet = workbook.active

            # Append data to the Excel sheet in a new row
            sheet.append([
                resume_data['name'],
                resume_data['email'],
                resume_data['mobile_number'],
                resume_data['linkedin_link'],
                resume_data['education'],
                resume_data['experience'],
                resume_data['skills'],
            ])

            workbook.save(file_path)  # Save changes to the existing workbook

            return render(request, 'success.html', {'resume_data': resume_data})
    else:
        form = ResumeUploadForm()

    return render(request, 'upload.html', {'form': form})
